from tkinter import *
from tkinter.ttk import Combobox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.geometry("990x500")
root.title("Data Entry Form")
root.maxsize(width=990,height=500)
root.minsize(width=990,height=500)
root.resizable(False,False)
root.config(bg="light blue")
bg_color="light blue"

file=pathlib.Path('data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Emp Id"
    sheet['B1']="Name"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Date of Birth"
    sheet['F1']="Category"
    sheet['G1']="Qualification"
    sheet['H1']="Religion"
    sheet['I1']="Nationality"
    sheet['J1']="City"
    sheet['K1']="Salary"
    sheet['L1']="Job Designation"
    
    file.save('data.xlsx')

def Submit():
    id=idvalue.get()
    name=namevalue.get()
    age=agevalue.get()
    gender=gender_combox.get()
    dob=dobvalue.get()
    category=category_combox.get()
    qualification=qualification_combox.get()
    religion=religion_combox.get()
    nationality=nationalityvalue.get()
    city=cityvalue.get()
    salary=salaryvalue.get()
    jobdesignations=jobdesignationsvalue.get()
    
    file=openpyxl.load_workbook('data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=id)
    sheet.cell(column=2,row=sheet.max_row,value=name)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=dob)
    sheet.cell(column=6,row=sheet.max_row,value=category)
    sheet.cell(column=7,row=sheet.max_row,value=qualification)
    sheet.cell(column=8,row=sheet.max_row,value=religion)
    sheet.cell(column=9,row=sheet.max_row,value=nationality)
    sheet.cell(column=10,row=sheet.max_row,value=city)
    sheet.cell(column=11,row=sheet.max_row,value=salary)
    sheet.cell(column=12,row=sheet.max_row,value=jobdesignations)
    
    file.save(r'data.xlsx')
    
def Clear():
    idvalue.set('')
    namevalue.set('')
    agevalue.set('')
    dobvalue.set('')
    nationalityvalue.set('')
    cityvalue.set('')
    salaryvalue.set('')
    jobdesignationsvalue.set('')

idvalue=StringVar()
namevalue=StringVar()
agevalue=StringVar()
dobvalue=StringVar()
nationalityvalue=StringVar()
cityvalue=StringVar()
salaryvalue=StringVar()
jobdesignationsvalue=StringVar()

id=Label(root,text="Employee ID",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
id.grid(row=0,column=0,sticky=W)
id_entry=Entry(root,textvariable=idvalue,font=("Comic Sans MS",12),bg=bg_color,bd=3,width=16)
id_entry.grid(row=0,column=2)

name=Label(root,text="Name",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
name.grid(row=1,column=0,sticky=W)
name_entry=Entry(root,textvariable=namevalue,font=("Comic Sans MS",12),bg=bg_color,bd=3,width=16)
name_entry.grid(row=1,column=2)

age=Label(root,text="Age",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
age.grid(row=2,column=0,sticky=W)
age_entry=Entry(root,textvariable=agevalue,font=("Comic Sans MS",12),bg=bg_color,bd=3,width=16)
age_entry.grid(row=2,column=2)

gender=Label(root,text="Gender",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
gender.grid(row=3,column=0,sticky=W)
gender_combox=Combobox(root,values=["Male","Female"],font=("Comic Sans MS",12),state='r',width=16,)
gender_combox.grid(row=3,column=2)
# gender_combox.set('male')

dob=Label(root,text="Date of Birth",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
dob.grid(row=4,column=0,sticky=W)
dob_entry=Entry(root,textvariable=dobvalue,font=("Comic Sans MS",12),bg=bg_color,bd=3,width=17)
dob_entry.grid(row=4,column=2)

category=Label(root,text="Category",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=16)
category.grid(row=5,column=0,sticky=W)
category_combox=Combobox(root,values=["General","OBC","SC/ST"],font=("Comic Sans MS",12),state='r',width=16,)
category_combox.grid(row=5,column=2)

qualification=Label(root,text="Qualification",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
qualification.grid(row=0,column=4,sticky=W)
qualification_combox=Combobox(root,values=["10th","12th","Graduation","Post Graduation"],font=("Comic Sans MS",12),state='r',width=16,)
qualification_combox.grid(row=0,column=5)

religion=Label(root,text="Religion",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
religion.grid(row=1,column=4,sticky=W)
religion_combox=Combobox(root,values=["Hindusim","Buddhism","Christianity","Islam"],font=("Comic Sans MS",12)
                        ,state='r',width=16,)
religion_combox.grid(row=1,column=5)
# religion_combox.set('')

nationality=Label(root,text="Nationality",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
nationality.grid(row=2,column=4,sticky=W)
nationality_entry=Entry(root,textvariable=nationalityvalue,font=("Comic Sans MS",12),bg=bg_color,bd=3,width=16)
nationality_entry.grid(row=2,column=5)

city=Label(root,text="City",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
city.grid(row=3,column=4,sticky=W)
city_entry=Entry(root,textvariable=cityvalue,font=("Comic Sans MS",12),bg=bg_color,bd=3,width=16)
city_entry.grid(row=3,column=5)

salary=Label(root,text="Salary",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
salary.grid(row=4,column=4,sticky=W)
salary_entry=Entry(root,textvariable=salaryvalue,font=("Comic Sans MS",12),bg=bg_color,bd=3,width=16)
salary_entry.grid(row=4,column=5)

jobdesignations=Label(root,text="Job Designation",font=("Comic Sans MS",18,"bold"),bg=bg_color,padx=50,pady=15)
jobdesignations.grid(row=5,column=4,sticky=W)
jobdesignations_entry=Entry(root,textvariable=jobdesignationsvalue,font=("Comic Sans MS",12),bg=bg_color,bd=3,width=16)
jobdesignations_entry.grid(row=5,column=5)

submit=Button(root,text="Submit",font=("Comic Sans MS",15,"bold"),bg=bg_color,bd=3,command=Submit)
submit.place(x=270,y=430,width=90,height=40)

clear=Button(root,text="Clear",font=("Comic Sans MS",15,"bold"),bg=bg_color,bd=3,command=Clear)
clear.place(x=390,y=430,width=90,height=40)

exit=Button(root,text="Exit",font=("Comic Sans MS",15,"bold"),bg=bg_color,bd=3,command=lambda:root.destroy())
exit.place(x=515,y=430,width=90,height=40)
root.mainloop()